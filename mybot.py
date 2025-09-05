import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from telegram import (
    Update,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    ReplyKeyboardRemove
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    CallbackQueryHandler,
)
import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import tempfile

BOT_TOKEN = "7757762485:AAHY5BrJ58YpdW50lAwRUsTwahtRDrd1RyA"
ADMIN_ID = 6550324099

user_state = {}

STATS_FILE = "stats.json"
users_data = {}

# ====== Persistent session for dbfather with retries & headers ======
dbfather_session = requests.Session()
dbfather_logged_in = False

def _init_dbfather_session():
    global dbfather_session
    dbfather_session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    })
    retry = Retry(
        total=3,
        backoff_factor=0.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
        raise_on_status=False
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    dbfather_session.mount("http://", adapter)
    dbfather_session.mount("https://", adapter)

_init_dbfather_session()

def _extract_csrf(html: str) -> str | None:
    soup = BeautifulSoup(html, "html.parser")
    token_input = soup.find("input", {"name": "csrf_token"})
    return token_input["value"] if token_input and token_input.has_attr("value") else None

def _is_login_page(html: str) -> bool:
    # heuristic: presence of login form with userId field and login button
    soup = BeautifulSoup(html, "html.parser")
    uid = soup.find("input", {"name": "userId"})
    login_btn = soup.find("button", {"name": "login"})
    return bool(uid and login_btn)

def login_dbfather(force=False) -> bool:
    """
    Logs in to dbfather. Returns True if logged-in; otherwise False.
    """
    global dbfather_logged_in

    if dbfather_logged_in and not force:
        return True

    try:
        login_url = "https://dbfather.42web.io/"
        # Step 1: GET login page to grab CSRF token
        resp_get = dbfather_session.get(login_url, timeout=15)
        if resp_get.status_code != 200:
            print(f"❌ Login GET failed, status {resp_get.status_code}")
            dbfather_logged_in = False
            return False

        csrf_token = _extract_csrf(resp_get.text)
        if not csrf_token:
            print("❌ Could not find csrf_token on login page")
            dbfather_logged_in = False
            return False

        # Step 2: POST login with token
        payload = {
            "userId": "1582832816",
            "force_login": "1",
            "login": "لاگ اِن",
            "csrf_token": csrf_token
        }
        headers = {
            "Origin": "https://dbfather.42web.io",
            "Referer": "https://dbfather.42web.io/",
        }
        resp_post = dbfather_session.post(login_url, data=payload, headers=headers, timeout=20, allow_redirects=True)

        if resp_post.status_code != 200:
            print(f"❌ Login POST failed, status {resp_post.status_code}")
            dbfather_logged_in = False
            return False

        # Sometimes the site redirects back to the same URL with logged-in view
        # Confirm by checking that we're no longer on the login form
        html = resp_post.text
        if _is_login_page(html):
            print("❌ Still on login page after POST")
            dbfather_logged_in = False
            return False

        dbfather_logged_in = True
        print("✅ Logged in successfully")
        return True

    except requests.exceptions.RequestException as e:
        print(f"Login network error: {e}")
        dbfather_logged_in = False
        return False
    except Exception as e:
        print(f"Login error: {e}")
        dbfather_logged_in = False
        return False

def premium_search(api: str, query: str) -> str:
    """
    Performs a premium search. Ensures logged-in, fetches CSRF for search form, posts search.
    Retries login once if bounced to login.
    """
    base_url = "https://dbfather.42web.io/"
    try:
        # Ensure logged in
        if not dbfather_logged_in:
            if not login_dbfather():
                return "❌ Could not login to premium source."

        # GET search page to extract CSRF for the form used to submit searches
        r_get = dbfather_session.get(base_url, timeout=20)
        if r_get.status_code != 200:
            return f"❌ Premium source not reachable (GET {r_get.status_code})."

        # If we got bounced back to login, re-login once
        if _is_login_page(r_get.text):
            if not login_dbfather(force=True):
                return "❌ Premium session expired and re-login failed."
            r_get = dbfather_session.get(base_url, timeout=20)
            if r_get.status_code != 200 or _is_login_page(r_get.text):
                return "❌ Could not reach premium search page after re-login."

        csrf_token = _extract_csrf(r_get.text)
        if not csrf_token:
            # Some pages may render results without token; but usually it's required
            print("⚠ No csrf_token found on search page; proceeding may fail.")
        
        # Build payload similar to site’s form
        payload = {
            "csrf_token": csrf_token or "",
            "api": api,                 # e.g., auto / cnic / vehicle / number / ptcl
            "searchQuery": query,
            "search": "1"               # mimic submit button (if needed)
        }
        headers = {
            "Origin": "https://dbfather.42web.io",
            "Referer": "https://dbfather.42web.io/",
        }
        r_post = dbfather_session.post(base_url, data=payload, headers=headers, timeout=25, allow_redirects=True)

        # If server closed connection early, requests (with retries) should have retried. If still failing:
        if r_post.status_code != 200:
            return f"❌ Premium search failed (POST {r_post.status_code})."

        # If we were bounced to login again, try one more re-login and one retry of POST
        if _is_login_page(r_post.text):
            if not login_dbfather(force=True):
                return "❌ Session lost during search; re-login failed."
            # Re-GET + extract csrf + re-POST
            r_get2 = dbfather_session.get(base_url, timeout=20)
            csrf_token2 = _extract_csrf(r_get2.text)
            payload["csrf_token"] = csrf_token2 or ""
            r_post = dbfather_session.post(base_url, data=payload, headers=headers, timeout=25, allow_redirects=True)
            if r_post.status_code != 200 or _is_login_page(r_post.text):
                return "❌ Premium search failed after re-login."

        # Parse results tables
        soup = BeautifulSoup(r_post.text, "html.parser")
        tables = soup.find_all("table")
        if not tables:
            # Try to fetch a known results container to give better feedback
            results_div = soup.find("div", {"id": "results"}) or soup.find("div", class_="results")
            if results_div:
                # Extract text in a readable way
                text = results_div.get_text("\n", strip=True)
                return text if text else "⚠ No result found."
            return "⚠ No result found."

        out_lines = []
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cols = row.find_all(["td", "th"])
                if not cols:
                    continue
                if len(cols) == 1:
                    out_lines.append(cols[0].get_text(strip=True))
                elif len(cols) >= 2:
                    key = cols[0].get_text(strip=True)
                    val = cols[1].get_text(strip=True)
                    out_lines.append(f"{key}: {val}")
            out_lines.append("")  # blank line between tables

        result_text = "\n".join(out_lines).strip()
        return result_text if result_text else "⚠ No result found."

    except requests.exceptions.RequestException as e:
        # Network-level issues including RemoteDisconnected after retries
        return f"❌ Premium network error: {e}"
    except Exception as e:
        return f"❌ Error: {e}"

# ====== Load stats ======
def load_stats():
    global users_data
    if not os.path.isfile(STATS_FILE):
        users_data = {}
        return
    try:
        with open(STATS_FILE, "r") as f:
            data = json.load(f)
            users_data = data if isinstance(data, dict) else {}
    except:
        users_data = {}

# ====== Save stats ======
def save_stats():
    try:
        with open(STATS_FILE, "w") as f:
            json.dump(users_data, f, indent=2)
    except:
        pass

# ====== Inline Keyboards ======
def get_main_inline_keyboard():
    keyboard = [
        [
            InlineKeyboardButton("🆓 Free Search", callback_data="free"),
            InlineKeyboardButton("💎 Premium Search", callback_data="premium"),
        ]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_free_inline_keyboard():
    keyboard = [
        [
            InlineKeyboardButton("📱 Search by Number", callback_data="search_number"),
            InlineKeyboardButton("🆔 Search by CNIC", callback_data="search_cnic"),
        ],
        [InlineKeyboardButton("⬅ Back", callback_data="back_main")]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_premium_inline_keyboard():
    keyboard = [
        [
            InlineKeyboardButton("🤖 Auto", callback_data="premium_auto"),
            InlineKeyboardButton("☎ PTCL Detail", callback_data="premium_ptcl"),
        ],
        [
            InlineKeyboardButton("📱 Number Ownership", callback_data="premium_number"),
            InlineKeyboardButton("🚗 Vehicle Detail", callback_data="premium_vehicle"),
        ],
        [
            InlineKeyboardButton("🆔 CNIC Detail", callback_data="premium_cnic"),
        ],
        [InlineKeyboardButton("⬅ Back", callback_data="back_main")]
    ]
    return InlineKeyboardMarkup(keyboard)

# ====== /start ======
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_id = str(user.id)
    username = user.username or user.first_name or "Unknown"

    if user_id not in users_data:
        users_data[user_id] = {"username": username, "search_count": 0, "searches": []}
        save_stats()
    else:
        users_data[user_id]["username"] = username
        save_stats()

    user_state.pop(update.effective_chat.id, None)
    await update.message.reply_text(
        "👋 Welcome! Please choose an option:",
        reply_markup=get_main_inline_keyboard()
    )

# ====== Callback Handler ======
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat_id

    # Free vs Premium
    if query.data == "free":
        await query.message.reply_text("🆓 Free Search Options:", reply_markup=get_free_inline_keyboard())
    elif query.data == "premium":
        await query.message.reply_text("💎 Premium Search Options:", reply_markup=get_premium_inline_keyboard())
    elif query.data == "back_main":
        await query.message.reply_text("⬅ Back to Main Menu:", reply_markup=get_main_inline_keyboard())

    # Free Search options
    elif query.data == "search_number":
        user_state[chat_id] = ("free", "number")
        await query.message.reply_text("📱 Please enter the mobile number (10 or 11 digits):", reply_markup=ReplyKeyboardRemove())
    elif query.data == "search_cnic":
        user_state[chat_id] = ("free", "cnic")
        await query.message.reply_text("🆔 Please enter the CNIC number (13 digits):", reply_markup=ReplyKeyboardRemove())

    # Premium Search options
    elif query.data.startswith("premium_"):
        api_type = query.data.replace("premium_", "")
        user_state[chat_id] = ("premium", api_type)
        await query.message.reply_text(f"💎 Please enter query for {api_type.title()} search:", reply_markup=ReplyKeyboardRemove())
    else:
        await query.message.reply_text("⚠ Unknown option.")

# ====== Handle searches ======
async def menu_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user = update.effective_user
    user_id = str(user.id)
    username = user.username or user.first_name or "Unknown"
    text = update.message.text.strip()

    if user_id not in users_data:
        users_data[user_id] = {"username": username, "search_count": 0, "searches": []}
    else:
        users_data[user_id]["username"] = username
    save_stats()

    if chat_id not in user_state:
        await update.message.reply_text("⚠ Please type /start and select an option.", reply_markup=get_main_inline_keyboard())
        return

    mode, search_type = user_state[chat_id]

    # ===== Free Search =====
    if mode == "free":
        if search_type == "number":
            if not text.isdigit() or len(text) not in [10, 11]:
                await update.message.reply_text("❌ Invalid number. Enter 10 or 11 digits.")
                return
        elif search_type == "cnic":
            if not text.isdigit() or len(text) != 13:
                await update.message.reply_text("❌ Invalid CNIC. Enter exactly 13 digits.")
                return

        users_data[user_id]["search_count"] += 1
        users_data[user_id]["searches"].append({"type": search_type, "query": text})
        save_stats()

        await update.message.reply_text("🔍 Searching (Free)... Please wait.")

        url = "https://minahalsimdata.com.pk/sim-info/"
        payload = {"searchinfo": text}
        try:
            response = requests.post(url, data=payload, timeout=20)
            soup = BeautifulSoup(response.text, "html.parser")
            result_containers = soup.find_all("div", class_="resultcontainer")

            if not result_containers:
                await update.message.reply_text("⚠ No result found.")
                await send_developer_info(update)
                return

            result_text = ""
            for container in result_containers:
                rows = container.find_all("div", class_="row")
                record = {}
                for row in rows:
                    head = row.find("span", class_="detailshead")
                    value = row.find("span", class_="details")
                    if head and value:
                        record[head.get_text(strip=True).replace(":", "")] = value.get_text(strip=True)

                result_text += (
                    f"👤 Name: {record.get('Name', 'N/A')}\n"
                    f"📱 Mobile: {record.get('Mobile', 'N/A')}\n"
                    f"🌍 Country: {record.get('Country', 'N/A')}\n"
                    f"🆔 CNIC: {record.get('CNIC', 'N/A')}\n"
                    f"🏠 Address: {record.get('Address', 'N/A')}\n\n"
                )

            await update.message.reply_text(result_text.strip() or "⚠ No data found.")
            await send_developer_info(update)

        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")
            await send_developer_info(update)

    # ===== Premium Search =====
    elif mode == "premium":
        users_data[user_id]["search_count"] += 1
        users_data[user_id]["searches"].append({"type": f"premium_{search_type}", "query": text})
        save_stats()

        await update.message.reply_text(f"🔍 Searching (Premium - {search_type.title()})...")

        try:
            result_text = premium_search(search_type, text)
            await update.message.reply_text(result_text.strip() or "⚠ No data found.")
            await send_developer_info(update)
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")
            await send_developer_info(update)

# ===== Developer info =====
async def send_developer_info(update: Update):
    developer_msg = "🤖 Bot developed by Muazam Ali\n📞 WhatsApp: "
    await update.message.reply_text(developer_msg)
    await update.message.reply_text("Choose search type:", reply_markup=get_main_inline_keyboard())
    user_state.pop(update.effective_chat.id, None)

# ====== Stats Command ======
async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Unauthorized.")
            return
        if not users_data:
            await update.message.reply_text("No users yet.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Stats"

        center = Alignment(horizontal="center", vertical="center")
        bold_font_white = Font(bold=True, color="FFFFFF")
        bold_font_black = Font(bold=True, color="000000")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        purple_fill = PatternFill(start_color="A47DB9", end_color="A47DB9", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        orange_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
        cyan_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

        row_num = 1

        for uid, data in users_data.items():
            username = data.get("username", "Unknown")
            search_count = data.get("search_count", 0)
            searches = data.get("searches", [])

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws.cell(row=row_num, column=1, value="User Name").fill = blue_fill
            ws.cell(row=row_num, column=1).alignment = center
            ws.cell(row=row_num, column=1).font = bold_font_white
            ws.cell(row=row_num, column=1).border = thin_border

            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=3, value="User Id").fill = purple_fill
            ws.cell(row=row_num, column=3).alignment = center
            ws.cell(row=row_num, column=3).font = bold_font_white
            ws.cell(row=row_num, column=3).border = thin_border

            row_num += 1

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws.cell(row=row_num, column=1, value=username).border = thin_border
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=3, value=uid).border = thin_border

            row_num += 1

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws.cell(row=row_num, column=1, value="Total Searches").fill = yellow_fill
            ws.cell(row=row_num, column=1).alignment = center
            ws.cell(row=row_num, column=1).font = bold_font_black
            ws.cell(row=row_num, column=1).border = thin_border

            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=3, value=search_count).fill = yellow_fill
            ws.cell(row=row_num, column=3).alignment = center
            ws.cell(row=row_num, column=3).font = bold_font_black
            ws.cell(row=row_num, column=3).border = thin_border

            row_num += 1

            ws.cell(row=row_num, column=1, value="SR").fill = green_fill
            ws.cell(row=row_num, column=2, value="Search Type").fill = orange_fill
            ws.cell(row=row_num, column=3, value="Search Query").fill = cyan_fill
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)

            for col in range(1, 5):
                ws.cell(row=row_num, column=col).alignment = center
                ws.cell(row=row_num, column=col).font = bold_font_white
                ws.cell(row=row_num, column=col).border = thin_border

            row_num += 1

            for idx, s in enumerate(searches, start=1):
                ws.cell(row=row_num, column=1, value=idx).fill = green_fill
                ws.cell(row=row_num, column=1).alignment = center

                ws.cell(row=row_num, column=2, value=s["type"]).fill = orange_fill
                ws.cell(row=row_num, column=2).alignment = center

                ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
                ws.cell(row=row_num, column=3, value=s["query"]).fill = cyan_fill
                ws.cell(row=row_num, column=3).alignment = center

                for col in range(1, 5):
                    ws.cell(row=row_num, column=col).border = thin_border

                row_num += 1

            row_num += 2

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)

        with open(tmp_path, "rb") as file:
            await context.bot.send_document(chat_id=ADMIN_ID, document=file, filename="user_stats.xlsx")

        os.remove(tmp_path)

    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")

# ====== Main ======
if __name__ == "__main__":
    load_stats()
    login_dbfather()  # login once at startup (will handle csrf)
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, menu_choice))
    print("🤖 Bot is running...")
    app.run_polling()
